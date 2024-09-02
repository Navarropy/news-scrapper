import requests
from lxml import html
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Define URLs and website names
websites = [
    {
        'url': 'https://edition.cnn.com/us',
        'website_name': 'CNN',
        'xpath': '//span[contains(@class, "headline")]/text()'
    },
    {
        'url': 'https://www.foxnews.com/us',
        'website_name': 'Fox News',
        'xpath': '//*[@class="title"]/descendant::a[not(parent::header)]/text()'
    }
]

# Initialize lists to hold all headlines
all_headlines = []

# Process each website
for site in websites:
    # Send a request to the website
    response = requests.get(site['url'])
    response.raise_for_status()  # Ensure we notice bad responses

    # Parse the content with lxml
    tree = html.fromstring(response.content)

    # Extract headlines using XPath
    headlines = tree.xpath(site['xpath'])

    # Filter out empty headlines or those that are a single word
    filtered_headlines = [headline.strip() for headline in headlines if len(headline.strip().split()) > 1]

    # Add to the list of all headlines with their source
    for headline in filtered_headlines:
        all_headlines.append({'headline': headline, 'website': site['website_name']})

# Connect to SQLite database (or create it if it doesn't exist)
conn = sqlite3.connect('headlines.db')
cursor = conn.cursor()

# Create a table for the headlines with a website column
cursor.execute('''
    CREATE TABLE IF NOT EXISTS headlines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        headline TEXT,
        website TEXT
    )
''')

# Insert all headlines into the database with the website information
for item in all_headlines:
    cursor.execute('INSERT INTO headlines (headline, website) VALUES (?, ?)', (item['headline'], item['website']))

# Commit changes and close the connection
conn.commit()
conn.close()

# Create a pandas DataFrame from all headlines and sort it by headline
df = pd.DataFrame(all_headlines).sort_values(by='headline')

# Export the DataFrame to an Excel spreadsheet
excel_path = 'headlines.xlsx'
df.to_excel(excel_path, index=False, engine='openpyxl')

# Load the workbook and get the active worksheet
wb = load_workbook(excel_path)
ws = wb.active

# Define a bold font style and center alignment
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center')

# Apply styles to the header row
for cell in ws[1]:
    cell.font = bold_font
    cell.alignment = center_align

# Adjust column widths for better readability
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 20

# Save the styled workbook
wb.save(excel_path)

print("Headlines from CNN and Fox News have been collected, ordered, stored in the database, and exported to 'headlines.xlsx' with improved formatting.")
